"""Multi-sheet Excel report exporter for reconciliation results."""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..reconciliation.engine import (
    ROW_MISMATCH,
    ROW_MISSING_IN_FILE_1,
    ROW_MISSING_IN_FILE_2,
    ReconciliationResult,
)
from ..reconciliation.value_compare import FORMAT_MISMATCH, VALUE_MISMATCH

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
LABEL_FONT = Font(bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
YELLOW_FONT = Font(color="9C6500")


def _autofit(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def _write_header_row(ws, row_idx: int, headers) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_summary_sheet(ws, result: ReconciliationResult, file_1_name: str, file_2_name: str) -> None:
    s = result.summary
    ws["A1"] = "Data Reconciliation Tool - Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = f"File 1: {file_1_name}    |    File 2: {file_2_name}"

    overall_badge = "MATCH" if s["overall_status"] == "MATCH" else "XMATCH"
    badge_cell = ws["A5"]
    badge_cell.value = f"Overall Status: {overall_badge}"
    badge_cell.font = Font(bold=True, size=12, color="FFFFFF")
    badge_cell.fill = GREEN_FILL if overall_badge == "MATCH" else RED_FILL
    ws.merge_cells("A5:B5")

    rows = [
        ("Total Records - File 1", s["total_records_file_1"]),
        ("Total Records - File 2", s["total_records_file_2"]),
        ("Mapped Fields", s["mapped_fields_count"]),
        ("Unmapped Fields (File 1)", s["unmapped_fields_file_1_count"]),
        ("Unmapped Fields (File 2)", s["unmapped_fields_file_2_count"]),
        ("Total Compared Records", s["total_compared_records"]),
        ("MATCH Records", s["match_count"]),
        ("MISMATCH Records", s["mismatch_count"]),
        ("Missing in File 1", s["missing_in_file_1_count"]),
        ("Missing in File 2", s["missing_in_file_2_count"]),
        ("Match Rate (%)", s["match_rate_percent"]),
        ("Duplicate Keys Dropped (File 1)", s["duplicate_keys_file_1"]),
        ("Duplicate Keys Dropped (File 2)", s["duplicate_keys_file_2"]),
    ]

    start_row = 7
    _write_header_row(ws, start_row, ["Metric", "Value"])
    for offset, (label, value) in enumerate(rows, start=1):
        r = start_row + offset
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        if label in ("MATCH Records",):
            cell.fill = GREEN_FILL
            cell.font = GREEN_FONT
        elif label in ("MISMATCH Records", "Missing in File 1", "Missing in File 2") and value:
            cell.fill = RED_FILL
            cell.font = RED_FONT

    _autofit(ws)


def _build_mismatch_log_sheet(ws, result: ReconciliationResult) -> None:
    headers = ["Primary Key", "Row Status", "Field", "File 1 Value", "File 2 Value", "Difference Type"]
    _write_header_row(ws, 1, headers)

    row_idx = 2
    df = result.detail_df
    for _, row in df.iterrows():
        row_status = row["Row_Status"]
        if row_status not in (ROW_MISMATCH, ROW_MISSING_IN_FILE_1, ROW_MISSING_IN_FILE_2):
            continue

        if row_status == ROW_MISMATCH:
            for label in result.field_labels:
                status = row.get(f"{label}::Status")
                if status in (FORMAT_MISMATCH, VALUE_MISMATCH):
                    ws.cell(row=row_idx, column=1, value=row["Primary Key"])
                    ws.cell(row=row_idx, column=2, value=row_status)
                    ws.cell(row=row_idx, column=3, value=label)
                    ws.cell(row=row_idx, column=4, value=row.get(f"{label}::File1"))
                    ws.cell(row=row_idx, column=5, value=row.get(f"{label}::File2"))
                    diff_cell = ws.cell(row=row_idx, column=6, value=status)
                    if status == VALUE_MISMATCH:
                        diff_cell.fill = RED_FILL
                        diff_cell.font = RED_FONT
                    else:
                        diff_cell.fill = YELLOW_FILL
                        diff_cell.font = YELLOW_FONT
                    row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value=row["Primary Key"])
            status_cell = ws.cell(row=row_idx, column=2, value=row_status)
            status_cell.fill = RED_FILL
            status_cell.font = RED_FONT
            ws.cell(row=row_idx, column=3, value="(entire record)")
            ws.cell(row=row_idx, column=4, value="")
            ws.cell(row=row_idx, column=5, value="")
            ws.cell(row=row_idx, column=6, value=row_status)
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No mismatches or missing records found. All data reconciled successfully.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    ws.freeze_panes = "A2"
    _autofit(ws)


def _build_unmapped_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Unmapped Fields - File 1 Only"
    ws["A1"].font = LABEL_FONT
    _write_header_row(ws, 2, ["Field Name (File 1)"])
    r = 3
    if result.unmapped_file_1:
        for field_name in result.unmapped_file_1:
            ws.cell(row=r, column=1, value=field_name)
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none - all File 1 fields are mapped)")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Unmapped Fields - File 2 Only").font = LABEL_FONT
    r += 1
    _write_header_row(ws, r, ["Field Name (File 2)"])
    r += 1
    if result.unmapped_file_2:
        for field_name in result.unmapped_file_2:
            ws.cell(row=r, column=1, value=field_name)
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none - all File 2 fields are mapped)")
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Mapped Fields (for reference)").font = LABEL_FONT
    r += 1
    _write_header_row(ws, r, ["File 1 Field", "File 2 Field", "Match Type", "Primary Key?"])
    r += 1
    pk_labels = set(result.primary_key_labels)
    for m in result.mapped_fields:
        from ..mapping.field_mapper import FieldMapping  # noqa: F401 (type context)

        label = m.field_1 if m.field_1 == m.field_2 else f"{m.field_1} / {m.field_2}"
        ws.cell(row=r, column=1, value=m.field_1)
        ws.cell(row=r, column=2, value=m.field_2)
        ws.cell(row=r, column=3, value=m.match_type)
        ws.cell(row=r, column=4, value="Yes" if label in pk_labels else "")
        r += 1

    _autofit(ws)


def export_to_excel(result: ReconciliationResult, file_1_name: str = "File 1", file_2_name: str = "File 2") -> io.BytesIO:
    """Build the full multi-sheet reconciliation report as an in-memory .xlsx file."""
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    _build_summary_sheet(ws_summary, result, file_1_name, file_2_name)

    ws_log = wb.create_sheet("Detailed Mismatch Log")
    _build_mismatch_log_sheet(ws_log, result)

    ws_unmapped = wb.create_sheet("Unmapped Fields Info")
    _build_unmapped_sheet(ws_unmapped, result)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
